#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def discover_project_root() -> Path:
    script_path = Path(__file__).resolve()
    for candidate in script_path.parents:
        if (candidate / "README.md").exists() and (candidate / "output").exists():
            return candidate
    return script_path.parents[1]


ROOT = discover_project_root()
DEFAULT_OUTPUT_DIR = ROOT / "output"
DEFAULT_SALES_XLSX = DEFAULT_OUTPUT_DIR / "销售数据_模拟.xlsx"
DEFAULT_AR_XLSX = DEFAULT_OUTPUT_DIR / "金域应收_模拟.xlsx"
DEFAULT_CHART_JS = DEFAULT_OUTPUT_DIR / "chart.umd.min.js"
DEFAULT_HTML_OUT = DEFAULT_OUTPUT_DIR / "demo_dashboard_v3.html"

SALES_SHEET = "销售数据源"
AR_SHEET = "金域应收"

SALES_REQUIRED_HEADERS = [
    "日期",
    "区域",
    "医院代码",
    "医院",
    "营销套餐",
    "标本量(单项)",
    "单项销售额 (折后)",
]

AR_REQUIRED_HEADERS = [
    "客户代码",
    "客户名称",
    "区域",
    "开票日期",
    "回款金额",
    "未回款金额",
    "账龄（天）",
]

SALES_HEADER_ALIASES = {
    "日期": ["送检日期", "业务日期"],
    "区域": ["合同区域名称", "销售区域"],
    "医院代码": ["客户代码", "医院编号"],
    "医院": ["客户名称", "医院名称"],
    "营销套餐": ["项目", "套餐", "项目套餐", "套餐名称", "营销项目"],
    "标本量(单项)": ["标本量", "标本数量"],
    "单项销售额 (折后)": ["折后销售额", "销售额(折后)", "销售额（折后）"],
}

AR_HEADER_ALIASES = {
    "客户代码": ["医院代码", "客户编号"],
    "客户名称": ["医院", "医院名称"],
    "区域": ["合同区域名称", "销售区域"],
    "开票日期": ["交易日期", "业务日期"],
    "回款金额": ["实收金额", "到账金额"],
    "未回款金额": ["应收余额", "未收金额"],
    "账龄（天）": ["账龄(天)", "账龄"],
}

SALES_OPTIONAL_HEADERS = {
    "city": ["地区", "城市", "地市", "销售城市"],
    "hospital_level": ["医院等级", "医院级别"],
    "subseries": ["亚系列", "项目子类"],
}

AR_OPTIONAL_HEADERS = {
    "invoice_amount": ["发票金额", "开票金额"],
    "payment_date": ["回款日期", "收款时间(核销)", "到账日期"],
    "credit_days": ["信用天数", "账期", "合同信用天数"],
    "city": ["地区", "城市", "地市", "销售城市"],
    "sales_area": ["销售小区名称", "区域小区", "销售片区"],
}


class ValidationError(Exception):
    pass


@dataclass
class DatasetMeta:
    row_count: int
    min_month: str
    max_month: str


@dataclass
class RuntimeConfig:
    sales_path: Path
    ar_path: Path
    chart_js_path: Path
    html_out_path: Path


def parse_args() -> RuntimeConfig:
    parser = argparse.ArgumentParser(
        description="Generate the offline weekly sales and collection dashboard."
    )
    parser.add_argument("--sales", type=Path, default=DEFAULT_SALES_XLSX, help="销售 xlsx 路径")
    parser.add_argument("--ar", type=Path, default=DEFAULT_AR_XLSX, help="回款 xlsx 路径")
    parser.add_argument(
        "--chart-js",
        type=Path,
        default=DEFAULT_CHART_JS,
        help="本地图表库路径，默认读取 output/chart.umd.min.js",
    )
    parser.add_argument(
        "--html-out",
        type=Path,
        default=DEFAULT_HTML_OUT,
        help="输出 HTML 路径",
    )
    args = parser.parse_args()
    return RuntimeConfig(
        sales_path=args.sales.expanduser().resolve(),
        ar_path=args.ar.expanduser().resolve(),
        chart_js_path=args.chart_js.expanduser().resolve(),
        html_out_path=args.html_out.expanduser().resolve(),
    )


def require_file(path: Path) -> None:
    if not path.exists():
        raise ValidationError(f"缺少输入文件: {path}")


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    return (
        text.replace("\n", "")
        .replace("\r", "")
        .replace(" ", "")
        .replace("（", "(")
        .replace("）", ")")
    )


def parse_date(raw: object) -> datetime:
    if isinstance(raw, datetime):
        return raw
    if raw is None or str(raw).strip() == "":
        raise ValidationError("发现空日期字段，无法生成看板。")
    text = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m", "%Y%m%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%Y-%m", "%Y/%m"):
                return parsed.replace(day=1)
            return parsed
        except ValueError:
            continue
    raise ValidationError(f"无法解析日期字段: {text}")


def parse_number(raw: object) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().replace(",", "")
    if not text:
        return 0.0
    return float(text)


def month_from_raw(raw: object) -> str:
    return parse_date(raw).strftime("%Y-%m")


def iso_date(raw: object) -> str:
    return parse_date(raw).strftime("%Y-%m-%d")


def month_key(month: str) -> tuple[int, int]:
    year, value = month.split("-")
    return int(year), int(value)


def choose_sheet(path: Path, expected_sheet: str, aliases: list[str], warnings: list[str]):
    require_file(path)
    workbook = load_workbook(path, data_only=True)
    for name in [expected_sheet, *aliases]:
        if name in workbook.sheetnames:
            if name != expected_sheet:
                warnings.append(f"{path.name}: 使用工作表 {name} 替代默认工作表 {expected_sheet}")
            return workbook[name]
    if len(workbook.sheetnames) == 1:
        warnings.append(
            f"{path.name}: 未找到默认工作表 {expected_sheet}，已回退到唯一工作表 {workbook.sheetnames[0]}"
        )
        return workbook[workbook.sheetnames[0]]
    raise ValidationError(
        f"{path.name} 缺少工作表 {expected_sheet}，当前仅有: {', '.join(workbook.sheetnames)}"
    )


def resolve_headers(
    headers: list[object],
    required: list[str],
    aliases: dict[str, list[str]],
    label: str,
) -> dict[str, int]:
    normalized = [normalize_header(item) for item in headers]
    lookup = {name: idx for idx, name in enumerate(normalized) if name}
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for target in required:
        candidates = [target, *aliases.get(target, [])]
        match = next((lookup[name] for name in map(normalize_header, candidates) if name in lookup), None)
        if match is None:
            missing.append(target)
        else:
            resolved[target] = match
    if missing:
        raise ValidationError(f"{label} 缺少关键列: {', '.join(missing)}")
    return resolved


def resolve_optional_headers(headers: list[object], aliases: dict[str, list[str]]) -> dict[str, int]:
    normalized = [normalize_header(item) for item in headers]
    lookup = {name: idx for idx, name in enumerate(normalized) if name}
    resolved: dict[str, int] = {}
    for target, candidates in aliases.items():
        match = next((lookup[name] for name in map(normalize_header, candidates) if name in lookup), None)
        if match is not None:
            resolved[target] = match
    return resolved


def cell_value(row: tuple[object, ...], idx: int | None, default: object = "") -> object:
    if idx is None or idx >= len(row):
        return default
    value = row[idx]
    return default if value is None else value


def read_sales_records(path: Path, warnings: list[str]) -> tuple[list[dict[str, object]], DatasetMeta]:
    sheet = choose_sheet(path, SALES_SHEET, ["销售", "销售数据", "Sheet1"], warnings)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValidationError(f"{path.name} 没有数据。")
    header = list(rows[0])
    required = resolve_headers(header, SALES_REQUIRED_HEADERS, SALES_HEADER_ALIASES, path.name)
    optional = resolve_optional_headers(header, SALES_OPTIONAL_HEADERS)

    records: list[dict[str, object]] = []
    months: list[str] = []
    for row in rows[1:]:
        raw_date = cell_value(row, required["日期"], "")
        if raw_date == "":
            continue
        month = month_from_raw(raw_date)
        months.append(month)
        project_name = str(cell_value(row, required["营销套餐"], "未命名项目") or "未命名项目")
        subseries = str(cell_value(row, optional.get("subseries"), "") or "")
        records.append(
            {
                "date": iso_date(raw_date),
                "month": month,
                "region": str(cell_value(row, required["区域"], "未分配")),
                "city": str(cell_value(row, optional.get("city"), "未分配")),
                "hospital_code": str(cell_value(row, required["医院代码"], "")),
                "hospital_name": str(cell_value(row, required["医院"], "未知医院")),
                "hospital_level": str(cell_value(row, optional.get("hospital_level"), "未知级别")),
                "project_name": project_name,
                "project_group": subseries or project_name,
                "specimen_qty": parse_number(cell_value(row, required["标本量(单项)"], 0)),
                "sales_post": parse_number(cell_value(row, required["单项销售额 (折后)"], 0)),
            }
        )
    if not records:
        raise ValidationError(f"{path.name} 没有可用销售记录。")
    months_sorted = sorted(set(months), key=month_key)
    return records, DatasetMeta(len(records), months_sorted[0], months_sorted[-1])


def read_ar_records(path: Path, warnings: list[str]) -> tuple[list[dict[str, object]], DatasetMeta]:
    sheet = choose_sheet(path, AR_SHEET, ["应收", "回款", "应收数据", "Sheet1"], warnings)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValidationError(f"{path.name} 没有数据。")
    header = list(rows[0])
    required = resolve_headers(header, AR_REQUIRED_HEADERS, AR_HEADER_ALIASES, path.name)
    optional = resolve_optional_headers(header, AR_OPTIONAL_HEADERS)

    records: list[dict[str, object]] = []
    months: list[str] = []
    for row in rows[1:]:
        raw_date = cell_value(row, required["开票日期"], "")
        if raw_date == "":
            continue
        month = month_from_raw(raw_date)
        months.append(month)
        payment_raw = cell_value(row, optional.get("payment_date"), "")
        payment_date = ""
        if payment_raw != "":
            try:
                payment_date = iso_date(payment_raw)
            except ValidationError:
                payment_date = ""
        records.append(
            {
                "invoice_date": iso_date(raw_date),
                "month": month,
                "region": str(cell_value(row, required["区域"], "未分配")),
                "city": str(cell_value(row, optional.get("city"), "")),
                "hospital_code": str(cell_value(row, required["客户代码"], "")),
                "hospital_name": str(cell_value(row, required["客户名称"], "未知医院")),
                "invoice_amount": parse_number(cell_value(row, optional.get("invoice_amount"), 0)),
                "received_amount": parse_number(cell_value(row, required["回款金额"], 0)),
                "unpaid_amount": parse_number(cell_value(row, required["未回款金额"], 0)),
                "aging_days": parse_number(cell_value(row, required["账龄（天）"], 0)),
                "credit_days": parse_number(cell_value(row, optional.get("credit_days"), 0)),
                "payment_date": payment_date,
                "sales_area": str(cell_value(row, optional.get("sales_area"), "")),
            }
        )
    if not records:
        raise ValidationError(f"{path.name} 没有可用回款记录。")
    months_sorted = sorted(set(months), key=month_key)
    return records, DatasetMeta(len(records), months_sorted[0], months_sorted[-1])


def annotate_ar_records(
    sales_records: list[dict[str, object]],
    ar_records: list[dict[str, object]],
) -> list[dict[str, object]]:
    profile_by_code: dict[str, dict[str, str]] = {}
    profile_by_name: dict[str, dict[str, str]] = {}
    for row in sales_records:
        profile = {
            "region": str(row["region"]),
            "city": str(row["city"]),
            "hospital_level": str(row["hospital_level"]),
        }
        code = str(row["hospital_code"])
        name = str(row["hospital_name"])
        if code and code not in profile_by_code:
            profile_by_code[code] = profile
        if name and name not in profile_by_name:
            profile_by_name[name] = profile
    annotated: list[dict[str, object]] = []
    for row in ar_records:
        profile = profile_by_code.get(str(row["hospital_code"])) or profile_by_name.get(str(row["hospital_name"])) or {}
        city = str(row["city"] or profile.get("city") or "未分配")
        region = str(row["region"] or profile.get("region") or "未分配")
        level = str(profile.get("hospital_level") or "未知级别")
        annotated.append(
            {
                **row,
                "city": city,
                "region": region,
                "hospital_level": level,
            }
        )
    return annotated


def build_payload(
    sales_records: list[dict[str, object]],
    ar_records: list[dict[str, object]],
    sales_meta: DatasetMeta,
    ar_meta: DatasetMeta,
    config: RuntimeConfig,
    warnings: list[str],
) -> dict[str, object]:
    months = sorted(
        {str(row["month"]) for row in sales_records} | {str(row["month"]) for row in ar_records},
        key=month_key,
    )
    if not months:
        raise ValidationError("未发现可用于看板展示的月份。")

    regions = sorted(
        {
            str(value)
            for value in [*(row["region"] for row in sales_records), *(row["region"] for row in ar_records)]
            if str(value).strip()
        }
    )
    cities = sorted(
        {
            str(value)
            for value in [*(row["city"] for row in sales_records), *(row["city"] for row in ar_records)]
            if str(value).strip()
        }
    )

    sales_total = sum(float(row["sales_post"]) for row in sales_records if row["month"] in months[-12:])
    unpaid_total = sum(float(row["unpaid_amount"]) for row in ar_records if row["month"] in months[-12:])
    received_total = sum(float(row["received_amount"]) for row in ar_records if row["month"] in months[-12:])
    invoice_total = sum(float(row["invoice_amount"]) for row in ar_records if row["month"] in months[-12:])

    return {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "generatedDate": datetime.now().strftime("%Y-%m-%d"),
        "months": months,
        "regions": regions,
        "cities": cities,
        "defaultRange": {
            "from": months[-12] if len(months) >= 12 else months[0],
            "to": months[-1],
        },
        "salesRecords": sales_records,
        "arRecords": ar_records,
        "meta": {
            "sales": sales_meta.__dict__,
            "ar": ar_meta.__dict__,
            "sourceFiles": {
                "sales": config.sales_path.name,
                "ar": config.ar_path.name,
            },
            "snapshot": {
                "sales_total": sales_total,
                "unpaid_total": unpaid_total,
                "collection_rate": (received_total / invoice_total) if invoice_total else 0.0,
            },
            "warnings": warnings,
        },
    }


def safe_script_text(text: str) -> str:
    return text.replace("</script", "<\\/script")


def render_html(payload: dict[str, object], chart_js_path: Path) -> str:
    require_file(chart_js_path)
    chart_js = safe_script_text(chart_js_path.read_text(encoding="utf-8"))
    payload_json = safe_script_text(json.dumps(payload, ensure_ascii=False))
    template = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>区域经理销售与回款周更看板</title>
  <style>
    :root{
      --bg:#f2f7f8;
      --ink:#12232b;
      --muted:#5a707d;
      --card:rgba(255,255,255,.9);
      --line:rgba(18,35,43,.08);
      --brand:#0f766e;
      --brand2:#2563eb;
      --warn:#d97706;
      --danger:#c2410c;
      --shadow:0 16px 40px rgba(18,35,43,.08);
      --radius:22px;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family:"Avenir Next","PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif;
      color:var(--ink);
      background:
        radial-gradient(circle at 0% 0%, rgba(15,118,110,.12), transparent 24%),
        radial-gradient(circle at 100% 10%, rgba(37,99,235,.10), transparent 20%),
        linear-gradient(180deg,#f9fcfc 0%, var(--bg) 52%, #f9fbfb 100%);
    }
    .shell{width:min(1440px,calc(100% - 24px));margin:0 auto;padding:20px 0 36px}
    .hero,.toolbar,.module,.kpi,.panel{
      background:var(--card);
      border:1px solid rgba(255,255,255,.75);
      box-shadow:var(--shadow);
      backdrop-filter:blur(8px);
    }
    .hero{
      border-radius:28px;
      padding:20px 22px;
      margin-bottom:14px;
      position:relative;
      overflow:hidden;
    }
    .hero:after{
      content:"";
      position:absolute;
      right:-70px;
      top:-60px;
      width:180px;
      height:180px;
      border-radius:50%;
      background:radial-gradient(circle, rgba(37,99,235,.14), transparent 68%);
      pointer-events:none;
    }
    .eyebrow{
      display:inline-flex;
      align-items:center;
      gap:8px;
      padding:7px 12px;
      border-radius:999px;
      background:rgba(15,118,110,.09);
      color:var(--brand);
      font-size:12px;
      font-weight:700;
      letter-spacing:.08em;
      text-transform:uppercase;
    }
    .hero-main{
      display:flex;
      align-items:flex-start;
      justify-content:space-between;
      gap:18px;
    }
    .hero-copy{max-width:920px}
    .hero h1{margin:12px 0 6px;font-size:clamp(28px,3.3vw,44px);line-height:1.02;letter-spacing:-.04em}
    .hero-summary{
      margin-top:14px;
      padding:12px 14px;
      border-radius:16px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.68);
      font-size:14px;
      line-height:1.6;
      color:var(--muted);
    }
    .hero-summary strong{color:var(--ink);font-size:15px}
    .meta-pills{
      display:flex;
      flex-wrap:wrap;
      gap:10px;
      justify-content:flex-end;
      position:relative;
      z-index:1;
    }
    .meta-pill{
      display:inline-flex;
      align-items:center;
      gap:8px;
      min-height:36px;
      padding:8px 12px;
      border-radius:14px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.74);
      color:var(--muted);
      font-size:12px;
    }
    .meta-pill strong{color:var(--ink);font-size:13px}
    .toolbar{
      border-radius:22px;
      padding:14px 16px;
      display:flex;
      flex-wrap:wrap;
      align-items:end;
      gap:12px;
      margin-bottom:14px;
    }
    .toolbar-title{font-size:13px;color:var(--muted);font-weight:700;letter-spacing:.06em;text-transform:uppercase;margin-right:6px}
    .field{display:flex;flex-direction:column;gap:6px;min-width:150px}
    .field label{font-size:12px;color:var(--muted);font-weight:600}
    .field select{
      padding:10px 12px;
      border-radius:12px;
      border:1px solid var(--line);
      background:#fff;
      color:var(--ink);
      font:inherit;
    }
    .grid{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:18px}
    .kpi{
      border-radius:22px;
      padding:16px 18px;
      min-height:138px;
      position:relative;
      overflow:hidden;
    }
    .kpi:after{
      content:"";
      position:absolute;
      right:-20px;
      bottom:-20px;
      width:108px;
      height:108px;
      border-radius:50%;
      background:radial-gradient(circle, rgba(37,99,235,.14), transparent 72%);
    }
    .kpi .label{font-size:13px;color:var(--muted);font-weight:600}
    .kpi .value{font-size:clamp(26px,3vw,38px);font-weight:800;letter-spacing:-.05em;margin:8px 0 6px}
    .kpi .hint{font-size:12px;color:var(--muted);line-height:1.55;max-width:26ch}
    .module{
      border-radius:26px;
      padding:18px;
      margin-top:16px;
    }
    .module-header{
      display:flex;
      justify-content:space-between;
      align-items:end;
      gap:16px;
      margin-bottom:12px;
    }
    .module-title{margin:4px 0 2px;font-size:clamp(22px,2vw,30px);letter-spacing:-.04em}
    .module-subtitle{margin:0;color:var(--muted);font-size:13px;line-height:1.5;max-width:72ch}
    .module-tag{
      padding:8px 12px;
      border-radius:999px;
      background:rgba(37,99,235,.08);
      color:var(--brand2);
      font-size:12px;
      font-weight:700;
      white-space:nowrap;
    }
    .highlights{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:16px}
    .highlight{
      flex:1 1 220px;
      padding:12px 14px;
      border-radius:18px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.74);
    }
    .highlight span{display:block;font-size:12px;color:var(--muted);margin-bottom:8px}
    .highlight strong{font-size:15px;line-height:1.55}
    .panels{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:18px}
    .panel{
      border-radius:22px;
      padding:16px;
      min-height:0;
    }
    .panel h3{margin:0;font-size:18px;letter-spacing:-.02em}
    .panel-note{margin:6px 0 12px;color:var(--muted);font-size:13px;line-height:1.55}
    .panel-meta{
      margin:0 0 12px;
      color:var(--ink);
      font-size:14px;
      font-weight:700;
      line-height:1.55;
    }
    .span-3{grid-column:span 3}.span-4{grid-column:span 4}.span-5{grid-column:span 5}
    .span-6{grid-column:span 6}.span-7{grid-column:span 7}.span-8{grid-column:span 8}
    .span-12{grid-column:span 12}
    .chart-wrap{
      position:relative;
      width:100%;
      min-height:220px;
    }
    .chart-wrap-xl{height:300px}
    .chart-wrap-lg{height:320px}
    .chart-wrap-md{height:300px}
    .chart-wrap-sm{height:240px}
    .chart-wrap-xs{height:210px}
    .chart-wrap canvas{
      display:block!important;
      width:100%!important;
      height:100%!important;
    }
    .panel-toolbar{
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:12px;
      margin-bottom:12px;
    }
    .panel-toolbar h3{margin:0}
    .panel-toolbar .field{
      min-width:220px;
      gap:4px;
    }
    .panel-toolbar .field select{padding:9px 11px}
    .table-wrap{
      overflow:auto;
      border-radius:16px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.65);
    }
    table{width:100%;border-collapse:collapse;min-width:820px}
    th,td{padding:12px 14px;text-align:left;font-size:13px;border-bottom:1px solid var(--line);vertical-align:top}
    thead th{
      position:sticky;top:0;z-index:1;
      background:#fff;
      font-size:12px;
      color:var(--muted);
      letter-spacing:.06em;
      text-transform:uppercase;
    }
    tbody tr:hover{background:rgba(15,118,110,.04)}
    .badge{
      display:inline-flex;
      align-items:center;
      padding:6px 10px;
      border-radius:999px;
      font-size:12px;
      font-weight:700;
    }
    .badge.high{background:rgba(194,65,12,.12);color:var(--danger)}
    .badge.medium{background:rgba(217,119,6,.12);color:var(--warn)}
    .badge.low{background:rgba(15,118,110,.12);color:var(--brand)}
    .detail-shell{
      display:grid;
      grid-template-columns:minmax(0,1fr) minmax(280px,36%);
      gap:16px;
      align-items:start;
    }
    .detail-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}
    .detail-box{
      border:1px solid var(--line);
      border-radius:16px;
      padding:12px 14px;
      background:rgba(255,255,255,.72);
    }
    .detail-box span{display:block;font-size:12px;color:var(--muted);margin-bottom:6px}
    .detail-box strong{font-size:16px;line-height:1.45}
    .detail-side{
      border:1px solid var(--line);
      border-radius:18px;
      padding:14px;
      background:rgba(255,255,255,.72);
    }
    .detail-side strong{display:block;font-size:14px;margin-bottom:8px}
    .detail-caption{
      margin-top:10px;
      color:var(--muted);
      font-size:12px;
      line-height:1.6;
    }
    .table-header{
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:12px;
      margin-bottom:12px;
    }
    .table-header h3{margin:0}
    .ghost-btn{
      appearance:none;
      border:1px solid var(--line);
      background:#fff;
      color:var(--ink);
      border-radius:999px;
      padding:8px 12px;
      font:inherit;
      font-size:12px;
      font-weight:700;
      cursor:pointer;
    }
    .table-summary{
      margin:0 0 12px;
      color:var(--muted);
      font-size:12px;
      line-height:1.55;
    }
    .tips{display:grid;gap:10px}
    .tip{
      padding:14px 16px;
      border-radius:16px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.76);
      font-size:14px;
      line-height:1.7;
    }
    .warn-box{
      margin-top:14px;
      padding:12px 14px;
      border-radius:16px;
      border:1px solid rgba(217,119,6,.18);
      background:rgba(217,119,6,.08);
      color:#8a4b0f;
      font-size:13px;
      line-height:1.7;
    }
    .fade-up{opacity:0;transform:translateY(18px);transition:opacity .45s ease, transform .45s ease}
    .fade-up.visible{opacity:1;transform:translateY(0)}
    @media (max-width:1200px){
      .hero-main{flex-direction:column}
      .meta-pills{justify-content:flex-start}
      .detail-shell{grid-template-columns:1fr}
      .span-8,.span-7,.span-6,.span-5,.span-4,.span-3{grid-column:span 12}
    }
    @media (max-width:720px){
      .shell{width:min(100% - 16px, 1440px);padding-top:14px}
      .hero,.module,.toolbar{padding:16px;border-radius:24px}
      .meta-pills,.highlights,.panel-toolbar,.table-header{flex-direction:column;align-items:stretch}
      .field{min-width:100%}
      .detail-grid{grid-template-columns:1fr}
      .table-wrap{border:0;background:transparent}
      table,thead,tbody,th,td,tr{display:block;width:100%}
      thead{display:none}
      table{min-width:0}
      tbody tr{
        margin-bottom:12px;
        border:1px solid var(--line);
        border-radius:16px;
        overflow:hidden;
        background:rgba(255,255,255,.88);
      }
      td{display:grid;gap:6px}
      td::before{
        content:attr(data-label);
        font-size:11px;
        color:var(--muted);
        letter-spacing:.06em;
        text-transform:uppercase;
      }
    }
  </style>
</head>
<body>
  <div class="shell">
    <section class="hero fade-up">
      <div class="hero-main">
        <div class="hero-copy">
          <div class="eyebrow">Weekly Operating Cockpit</div>
          <h1>区域经理销售与回款周更看板</h1>
        </div>
        <div class="meta-pills">
          <div class="meta-pill">默认窗口<strong id="meta-range">-</strong></div>
          <div class="meta-pill">刷新时间<strong id="meta-generated">-</strong></div>
        </div>
      </div>
      <div class="hero-summary">经营摘要：<strong id="summary-text">正在生成经营摘要...</strong></div>
      <div id="warning-box" class="warn-box" style="display:none"></div>
    </section>

    <section class="toolbar fade-up">
      <div class="toolbar-title">筛选控制</div>
      <div class="field">
        <label for="from-month">开始月份</label>
        <select id="from-month"></select>
      </div>
      <div class="field">
        <label for="to-month">结束月份</label>
        <select id="to-month"></select>
      </div>
      <div class="field">
        <label for="region-filter">区域</label>
        <select id="region-filter"></select>
      </div>
      <div class="field">
        <label for="city-filter">城市</label>
        <select id="city-filter"></select>
      </div>
    </section>

    <section class="grid fade-up">
      <article class="kpi span-3">
        <div class="label">折后销售额</div>
        <div class="value" id="kpi-sales">-</div>
        <div class="hint" id="kpi-sales-hint">观察时间窗口内的整体销售规模与结构稳定性。</div>
      </article>
      <article class="kpi span-3">
        <div class="label">标本量</div>
        <div class="value" id="kpi-specimen">-</div>
        <div class="hint" id="kpi-specimen-hint">反映业务活跃度和临床上量情况。</div>
      </article>
      <article class="kpi span-3">
        <div class="label">未回款金额</div>
        <div class="value" id="kpi-unpaid">-</div>
        <div class="hint" id="kpi-unpaid-hint">用于识别欠款规模和逾期压力。</div>
      </article>
      <article class="kpi span-3">
        <div class="label">回款率</div>
        <div class="value" id="kpi-rate">-</div>
        <div class="hint" id="kpi-rate-hint">基于当前筛选范围内的发票与回款记录实时计算。</div>
      </article>
    </section>

    <section class="module fade-up">
      <div class="module-header">
        <div>
          <div class="eyebrow">Module 01</div>
          <h2 class="module-title">销售驾驶舱</h2>
          <p class="module-subtitle">首屏先回答两个问题：最近 12 个月卖得怎么样，以及当前主要由哪些医院拉动。</p>
        </div>
        <div class="module-tag" id="sales-tag">最近 12 个月 / 全部区域</div>
      </div>
      <div class="panels">
        <article class="panel span-8">
          <h3>月度经营趋势｜观察销售与回款同步表现</h3>
          <p class="panel-note">柱看折后销售额，线看月度回款率，优先识别节奏变化和回款承压月份。</p>
          <p class="panel-meta" id="sales-insight">正在生成销售判断...</p>
          <div class="chart-wrap chart-wrap-xl">
            <canvas id="chart-monthly"></canvas>
          </div>
        </article>
        <article class="panel span-4">
          <h3>销量前十医院｜锁定核心贡献医院</h3>
          <p class="panel-note">保持降序排序，并直接标出金额，方便每天看主力医院是否稳盘。</p>
          <p class="panel-meta" id="hospital-insight">正在识别销量贡献最高的医院...</p>
          <div class="chart-wrap chart-wrap-md">
            <canvas id="chart-hospital-top10"></canvas>
          </div>
        </article>
      </div>
    </section>

    <section class="module fade-up">
      <div class="module-header">
        <div>
          <div class="eyebrow">Module 02</div>
          <h2 class="module-title">销售结构复盘</h2>
          <p class="module-subtitle">第二层用项目结构和三级医院趋势说明增长来自哪里，以及是否具备持续性。</p>
        </div>
        <div class="module-tag">结构贡献 / 趋势复盘</div>
      </div>
      <div class="panels">
        <article class="panel span-4">
          <h3>销量前十项目｜判断结构是否集中</h3>
          <p class="panel-note">适合月底复盘时看项目结构变化，以及本月还应继续推动哪些套餐。</p>
          <p class="panel-meta" id="project-insight">正在识别核心项目结构...</p>
          <div class="chart-wrap chart-wrap-md">
            <canvas id="chart-project-top10"></canvas>
          </div>
        </article>
        <article class="panel span-8">
          <h3>三级医院 3 年趋势｜看结构增长是否稳定</h3>
          <p class="panel-note">只保留季度级观察节奏，并在标题下给出最新值与同比变化。</p>
          <p class="panel-meta" id="tier3-summary">正在计算三级医院趋势摘要...</p>
          <div class="chart-wrap chart-wrap-lg">
            <canvas id="chart-tier3-trend"></canvas>
          </div>
        </article>
      </div>
    </section>

    <section class="module fade-up">
      <div class="module-header">
        <div>
          <div class="eyebrow">Module 03</div>
          <h2 class="module-title">回款分析</h2>
          <p class="module-subtitle">先看整体风险，再锁定重点医院和单院回款节奏，最后再看完整预测表明细。</p>
        </div>
        <div class="module-tag" id="collection-tag">规则预测 / 业务参考</div>
      </div>
      <div class="highlights">
        <div class="highlight">
          <span>回款判断</span>
          <strong id="collection-insight">正在计算回款总体表现...</strong>
        </div>
        <div class="highlight">
          <span>预测重点</span>
          <strong id="forecast-insight">正在识别需要优先跟进的回款对象...</strong>
        </div>
        <div class="highlight">
          <span>逾期结构</span>
          <strong id="aging-insight">正在识别账龄结构和超信用期情况...</strong>
        </div>
      </div>
      <div class="panels">
        <article class="panel span-6">
          <h3>账龄结构｜比较不同账龄段金额压力</h3>
          <p class="panel-note">按账龄从低到高排序展示金额，更适合比较欠款结构，而不是只看占比。</p>
          <div class="chart-wrap chart-wrap-sm">
            <canvas id="chart-aging"></canvas>
          </div>
        </article>
        <article class="panel span-6">
          <h3>重点风险医院 Top 5｜先看最该催回的对象</h3>
          <p class="panel-note">只保留最需要优先行动的 5 家医院，用于每天盯催和月底复盘。</p>
          <div class="table-wrap">
            <table id="risk-table">
              <thead>
                <tr>
                  <th>医院</th>
                  <th>未回款</th>
                  <th>最大账龄</th>
                  <th>风险</th>
                </tr>
              </thead>
              <tbody></tbody>
            </table>
          </div>
        </article>
        <article class="panel span-12">
          <div class="panel-toolbar">
            <div>
              <h3>某医院回款周期｜看单院回款节奏是否失控</h3>
              <p class="panel-note">医院选择器只影响这个分析面板，不改变整张预测表的排序。</p>
            </div>
            <div class="field">
              <label for="hospital-filter">选择医院</label>
              <select id="hospital-filter"></select>
            </div>
          </div>
          <div class="detail-shell">
            <div>
              <div class="detail-grid" id="hospital-detail-grid"></div>
            </div>
            <div class="detail-side">
              <strong>最近 6 次回款周期</strong>
              <div class="chart-wrap chart-wrap-xs">
                <canvas id="chart-hospital-cycle"></canvas>
              </div>
              <div class="detail-caption" id="hospital-cycle-caption">正在生成该医院的回款周期摘要...</div>
            </div>
          </div>
        </article>
        <article class="panel span-12">
          <div class="table-header">
            <div>
              <h3>医院回款预测表｜作为完整明细层查看</h3>
              <p class="panel-note">规则预测口径：历史回款周期、最近回款周期和信用天数加权，超信用期和长账龄医院自动上调风险。</p>
            </div>
            <button id="forecast-toggle" type="button" class="ghost-btn" style="display:none">展开全部</button>
          </div>
          <p class="table-summary" id="forecast-table-summary">当前默认展示前 8 家医院。</p>
          <div class="table-wrap">
            <table id="forecast-table">
              <thead>
                <tr>
                  <th>医院</th>
                  <th>区域 / 城市</th>
                  <th>未回款</th>
                  <th>平均回款周期</th>
                  <th>最近回款周期</th>
                  <th>信用天数</th>
                  <th>预计回款日期</th>
                  <th>风险等级</th>
                  <th>建议动作</th>
                </tr>
              </thead>
              <tbody></tbody>
            </table>
          </div>
        </article>
      </div>
    </section>

    <section class="module fade-up">
      <div class="module-header">
        <div>
          <div class="eyebrow">Module 04</div>
          <h2 class="module-title">轻量计划提示</h2>
          <p class="module-subtitle">不做完整任务流，只根据当前分析结果生成 3-5 条本周 / 本月动作建议，方便你快速落计划。</p>
        </div>
        <div class="module-tag">行动建议 / 自动生成</div>
      </div>
      <div class="tips" id="tips-list"></div>
    </section>
  </div>

  <script>__CHART_JS__</script>
  <script id="dashboard-data" type="application/json">__DATA_JSON__</script>
  <script>
    const RAW = JSON.parse(document.getElementById('dashboard-data').textContent);
    const state = {
      from: RAW.defaultRange.from,
      to: RAW.defaultRange.to,
      region: '全部',
      city: '全部',
      hospital: '',
      forecastExpanded: false,
    };
    const charts = {};
    const agingBuckets = ['0-30 天', '31-60 天', '61-90 天', '91-180 天', '181+ 天'];

    function formatNumber(value, digits = 0) {
      return new Intl.NumberFormat('zh-CN', {
        minimumFractionDigits: digits,
        maximumFractionDigits: digits,
      }).format(Number(value || 0));
    }

    function formatCompact(value, suffix = '') {
      const num = Number(value || 0);
      if (Math.abs(num) >= 100000000) return `${(num / 100000000).toFixed(2)}亿${suffix}`;
      if (Math.abs(num) >= 10000) return `${(num / 10000).toFixed(1)}万${suffix}`;
      return `${formatNumber(num)}${suffix}`;
    }

    function formatCurrency(value) {
      return `${(Number(value || 0) / 10000).toFixed(1)}万元`;
    }

    function formatPercent(value) {
      return `${(Number(value || 0) * 100).toFixed(1)}%`;
    }

    function formatPercentAxis(value) {
      return `${Number(value || 0).toFixed(0)}%`;
    }

    function truncateLabel(label, maxLength = 8) {
      const text = String(label || '');
      return text.length > maxLength ? `${text.slice(0, maxLength)}…` : text;
    }

    function categoryTickLabel(scale, value, maxLength = 10) {
      if (!scale || typeof scale.getLabelForValue !== 'function') return '';
      return truncateLabel(scale.getLabelForValue(value), maxLength);
    }

    function quarterTickLabel(scale, value) {
      if (!scale || typeof scale.getLabelForValue !== 'function') return '';
      const label = scale.getLabelForValue(value);
      const month = Number(String(label).slice(5, 7));
      return [3, 6, 9, 12].includes(month) ? label : '';
    }

    function parseDate(dateText) {
      if (!dateText) return null;
      return new Date(`${dateText}T00:00:00`);
    }

    function isoDate(date) {
      if (!date) return '';
      const year = date.getFullYear();
      const month = `${date.getMonth() + 1}`.padStart(2, '0');
      const day = `${date.getDate()}`.padStart(2, '0');
      return `${year}-${month}-${day}`;
    }

    function addDays(dateText, days) {
      const base = parseDate(dateText) || parseDate(RAW.generatedDate);
      base.setDate(base.getDate() + Math.max(0, Math.round(days || 0)));
      return isoDate(base);
    }

    function diffDays(startText, endText) {
      const start = parseDate(startText);
      const end = parseDate(endText);
      if (!start || !end) return null;
      const days = Math.round((end.getTime() - start.getTime()) / 86400000);
      return days >= 0 ? days : null;
    }

    function animateValue(el, formatter, target) {
      const finalValue = Number(target || 0);
      const duration = 820;
      const start = performance.now();
      const frame = (now) => {
        const progress = Math.min(1, (now - start) / duration);
        const eased = 1 - Math.pow(1 - progress, 3);
        el.textContent = formatter(finalValue * eased);
        if (progress < 1) requestAnimationFrame(frame);
      };
      requestAnimationFrame(frame);
    }

    function matchesRegion(row) {
      return state.region === '全部' || row.region === state.region;
    }

    function matchesCity(row) {
      return state.city === '全部' || row.city === state.city;
    }

    function matchesTime(row) {
      return row.month >= state.from && row.month <= state.to;
    }

    function bucketAging(days) {
      if (days <= 30) return '0-30 天';
      if (days <= 60) return '31-60 天';
      if (days <= 90) return '61-90 天';
      if (days <= 180) return '91-180 天';
      return '181+ 天';
    }

    function resolveRiskLevel(entry) {
      let score = 0;
      const overdueShare = entry.unpaid > 0 ? entry.overCreditAmount / entry.unpaid : 0;
      if (entry.unpaid >= 1500000) score += 2;
      if (entry.maxAging >= 180) score += 2;
      if (entry.avgCycle && entry.avgCycle > entry.creditDays + 20) score += 1;
      if (entry.weightedAging > entry.creditDays) score += 1;
      if (overdueShare >= 0.45) score += 2;
      if (score >= 5) return '高风险';
      if (score >= 3) return '中风险';
      return '常规跟进';
    }

    function badgeClass(level) {
      if (level === '高风险') return 'high';
      if (level === '中风险') return 'medium';
      return 'low';
    }

    function suggestionText(entry) {
      if (entry.riskLevel === '高风险') {
        return '优先跟进财务与科室负责人，明确回款节点并周内复盘';
      }
      if (entry.riskLevel === '中风险') {
        return '保持节奏跟进，补齐资料并锁定预计回款时间';
      }
      return '常规维护回款节奏，避免账龄继续上升';
    }

    function computeForecastCycle(entry) {
      const credit = entry.creditDays || 60;
      const avgCycle = entry.avgCycle || 0;
      const latestCycle = entry.latestCycle || 0;
      let cycle;
      if (entry.cycles.length >= 3) {
        cycle = 0.45 * avgCycle + 0.35 * latestCycle + 0.2 * credit;
      } else if (entry.cycles.length >= 1) {
        cycle = 0.55 * latestCycle + 0.45 * credit;
      } else {
        cycle = credit + 15;
      }
      if (entry.weightedAging > credit) {
        cycle = Math.max(cycle, entry.weightedAging + 10);
      }
      if (entry.maxAging > 180) {
        cycle += 10;
      }
      return Math.round(cycle);
    }

    function buildAggregate() {
      const filteredSales = RAW.salesRecords.filter((row) => matchesTime(row) && matchesRegion(row) && matchesCity(row));
      const filteredAr = RAW.arRecords.filter((row) => matchesTime(row) && matchesRegion(row) && matchesCity(row));
      const months = RAW.months.filter((month) => month >= state.from && month <= state.to);

      const monthlyMap = Object.fromEntries(months.map((month) => [month, 0]));
      const monthlyArMap = Object.fromEntries(months.map((month) => [month, { invoice: 0, received: 0 }]));
      const hospitalSalesMap = new Map();
      const projectMap = new Map();
      const citySalesMap = new Map();

      filteredSales.forEach((row) => {
        if (!monthlyMap[row.month] && monthlyMap[row.month] !== 0) {
          monthlyMap[row.month] = 0;
        }
        monthlyMap[row.month] += row.sales_post;
        citySalesMap.set(row.city, (citySalesMap.get(row.city) || 0) + row.sales_post);
        projectMap.set(row.project_name, (projectMap.get(row.project_name) || 0) + row.sales_post);
        const key = row.hospital_code || row.hospital_name;
        const entry = hospitalSalesMap.get(key) || {
          hospital: row.hospital_name,
          region: row.region,
          city: row.city,
          level: row.hospital_level,
          sales: 0,
          specimen: 0,
        };
        entry.sales += row.sales_post;
        entry.specimen += row.specimen_qty;
        hospitalSalesMap.set(key, entry);
      });

      const hospitalCollectionMap = new Map();
      const agingMap = Object.fromEntries(agingBuckets.map((label) => [label, 0]));
      filteredAr.forEach((row) => {
        if (!monthlyArMap[row.month]) {
          monthlyArMap[row.month] = { invoice: 0, received: 0 };
        }
        monthlyArMap[row.month].invoice += row.invoice_amount;
        monthlyArMap[row.month].received += row.received_amount;

        const key = row.hospital_code || row.hospital_name;
        const entry = hospitalCollectionMap.get(key) || {
          hospital: row.hospital_name,
          region: row.region,
          city: row.city,
          level: row.hospital_level,
          invoice: 0,
          received: 0,
          unpaid: 0,
          maxAging: 0,
          overCreditAmount: 0,
          agingWeight: 0,
          agingWeightedSum: 0,
          creditSamples: [],
          cycles: [],
          recentCycleItems: [],
          lastInvoiceDate: '',
          outstandingInvoiceDate: '',
          sales: 0,
          specimen: 0,
        };

        entry.invoice += row.invoice_amount;
        entry.received += row.received_amount;
        entry.unpaid += row.unpaid_amount;
        entry.maxAging = Math.max(entry.maxAging, row.aging_days);
        if (row.credit_days > 0) {
          entry.creditSamples.push(row.credit_days);
          if (row.aging_days > row.credit_days) {
            entry.overCreditAmount += row.unpaid_amount;
          }
        }

        const weight = Math.max(row.unpaid_amount || row.invoice_amount || 1, 1);
        entry.agingWeight += weight;
        entry.agingWeightedSum += row.aging_days * weight;
        entry.lastInvoiceDate = !entry.lastInvoiceDate || row.invoice_date > entry.lastInvoiceDate ? row.invoice_date : entry.lastInvoiceDate;
        if (row.unpaid_amount > 0) {
          entry.outstandingInvoiceDate = !entry.outstandingInvoiceDate || row.invoice_date > entry.outstandingInvoiceDate ? row.invoice_date : entry.outstandingInvoiceDate;
        }

        const cycle = diffDays(row.invoice_date, row.payment_date);
        if (cycle !== null && row.received_amount > 0) {
          entry.cycles.push(cycle);
          entry.recentCycleItems.push({
            date: row.payment_date,
            cycle,
            received: row.received_amount,
          });
        }

        agingMap[bucketAging(row.aging_days)] += row.unpaid_amount;
        hospitalCollectionMap.set(key, entry);
      });

      hospitalSalesMap.forEach((salesEntry, key) => {
        const collectionEntry = hospitalCollectionMap.get(key) || {
          hospital: salesEntry.hospital,
          region: salesEntry.region,
          city: salesEntry.city,
          level: salesEntry.level,
          invoice: 0,
          received: 0,
          unpaid: 0,
          maxAging: 0,
          overCreditAmount: 0,
          agingWeight: 0,
          agingWeightedSum: 0,
          creditSamples: [],
          cycles: [],
          recentCycleItems: [],
          lastInvoiceDate: '',
          outstandingInvoiceDate: '',
          sales: 0,
          specimen: 0,
        };
        collectionEntry.sales += salesEntry.sales;
        collectionEntry.specimen += salesEntry.specimen;
        collectionEntry.level = salesEntry.level || collectionEntry.level;
        collectionEntry.city = salesEntry.city || collectionEntry.city;
        collectionEntry.region = salesEntry.region || collectionEntry.region;
        hospitalCollectionMap.set(key, collectionEntry);
      });

      const hospitals = [...hospitalCollectionMap.values()].map((entry) => {
        entry.creditDays = entry.creditSamples.length
          ? entry.creditSamples.reduce((sum, value) => sum + value, 0) / entry.creditSamples.length
          : 60;
        entry.weightedAging = entry.agingWeight ? entry.agingWeightedSum / entry.agingWeight : 0;
        entry.avgCycle = entry.cycles.length
          ? entry.cycles.reduce((sum, value) => sum + value, 0) / entry.cycles.length
          : 0;
        entry.recentCycleItems.sort((a, b) => (a.date < b.date ? 1 : -1));
        entry.latestCycle = entry.recentCycleItems.length ? entry.recentCycleItems[0].cycle : 0;
        entry.collectionRate = entry.invoice ? entry.received / entry.invoice : 0;
        entry.forecastCycle = computeForecastCycle(entry);
        const baseDate = entry.outstandingInvoiceDate || entry.lastInvoiceDate || RAW.generatedDate;
        entry.predictedDate = addDays(baseDate, entry.forecastCycle);
        entry.riskLevel = resolveRiskLevel(entry);
        entry.suggestion = suggestionText(entry);
        return entry;
      });

      hospitals.sort((a, b) => {
        const riskRank = { '高风险': 3, '中风险': 2, '常规跟进': 1 };
        if (riskRank[b.riskLevel] !== riskRank[a.riskLevel]) return riskRank[b.riskLevel] - riskRank[a.riskLevel];
        if (b.unpaid !== a.unpaid) return b.unpaid - a.unpaid;
        return String(a.predictedDate || '9999-12-31').localeCompare(String(b.predictedDate || '9999-12-31'));
      });

      const topHospitals = [...hospitalSalesMap.values()]
        .sort((a, b) => b.sales - a.sales)
        .slice(0, 10);

      const topProjects = [...projectMap.entries()]
        .map(([name, value]) => ({ name, value }))
        .sort((a, b) => b.value - a.value)
        .slice(0, 10);

      const tier3Sales = RAW.salesRecords
        .filter((row) => row.month <= state.to && matchesRegion(row) && matchesCity(row) && String(row.hospital_level || '').includes('三'))
        .reduce((acc, row) => {
          acc[row.month] = (acc[row.month] || 0) + row.sales_post;
          return acc;
        }, {});
      const tier3Months = Object.keys(tier3Sales).sort().slice(-36);
      const tier3Trend = tier3Months.map((month) => ({ month, value: tier3Sales[month] || 0 }));
      const tier3Latest = tier3Trend.length ? tier3Trend[tier3Trend.length - 1] : null;
      const tier3LastYear = tier3Trend.length > 12 ? tier3Trend[tier3Trend.length - 13] : null;
      const tier3YoY = tier3Latest && tier3LastYear && tier3LastYear.value
        ? (tier3Latest.value - tier3LastYear.value) / tier3LastYear.value
        : null;

      const hospitalOptions = hospitals
        .filter((item) => item.sales > 0 || item.unpaid > 0)
        .sort((a, b) => (b.sales + b.unpaid) - (a.sales + a.unpaid))
        .map((item) => item.hospital);

      if (!hospitalOptions.length) {
        state.hospital = '';
      } else if (!state.hospital || !hospitalOptions.includes(state.hospital)) {
        state.hospital = hospitalOptions[0];
      }

      const selectedHospital = hospitals.find((item) => item.hospital === state.hospital) || hospitals[0] || null;
      const selectedHospitalCycles = selectedHospital
        ? selectedHospital.recentCycleItems
          .slice(0, 6)
          .reverse()
          .map((item) => ({
            label: item.date ? item.date.slice(5) : '未知日期',
            value: item.cycle,
            received: item.received,
          }))
        : [];
      const salesTotal = filteredSales.reduce((sum, row) => sum + row.sales_post, 0);
      const specimenTotal = filteredSales.reduce((sum, row) => sum + row.specimen_qty, 0);
      const unpaidTotal = filteredAr.reduce((sum, row) => sum + row.unpaid_amount, 0);
      const receivedTotal = filteredAr.reduce((sum, row) => sum + row.received_amount, 0);
      const invoiceTotal = filteredAr.reduce((sum, row) => sum + row.invoice_amount, 0);
      const collectionRate = invoiceTotal ? receivedTotal / invoiceTotal : 0;
      const overdueAmount = hospitals.reduce((sum, item) => sum + item.overCreditAmount, 0);
      const overdueShare = unpaidTotal ? overdueAmount / unpaidTotal : 0;
      const topHospital = topHospitals[0];
      const topProject = topProjects[0];
      const strongestCity = [...citySalesMap.entries()].sort((a, b) => b[1] - a[1])[0];
      const topRisk = hospitals[0];

      const planTips = [];
      if (topHospital) {
        planTips.push(`先稳住 ${topHospital.hospital}，当前销量 ${formatCurrency(topHospital.sales)}，建议优先做续量和结构提升。`);
      }
      if (topRisk && topRisk.unpaid > 0) {
        planTips.push(`${topRisk.hospital} 是当前最需要推动回款的医院，未回款 ${formatCurrency(topRisk.unpaid)}，建议本周明确回款节点。`);
      }
      if (topProject) {
        planTips.push(`本月重点项目可继续围绕 ${topProject.name} 推进，当前贡献 ${formatCurrency(topProject.value)}。`);
      }
      if (overdueShare > 0.35) {
        planTips.push(`超信用期欠款占比已达 ${formatPercent(overdueShare)}，月底前应单独拉出逾期医院做回款跟进表。`);
      }
      if (strongestCity) {
        planTips.push(`${strongestCity[0]} 当前是主要销售贡献城市，建议同步关注该城市的上量稳定性与回款节奏。`);
      }

      return {
        metrics: { salesTotal, specimenTotal, unpaidTotal, collectionRate, overdueShare },
        months,
        monthlySales: months.map((month) => ({ month, value: monthlyMap[month] || 0 })),
        monthlyCollectionRate: months.map((month) => {
          const snapshot = monthlyArMap[month] || { invoice: 0, received: 0 };
          return {
            month,
            value: snapshot.invoice ? snapshot.received / snapshot.invoice : 0,
          };
        }),
        topHospitals,
        topProjects,
        tier3Trend,
        tier3Summary: tier3Latest
          ? `${tier3Latest.month} 折后销量 ${formatCurrency(tier3Latest.value)}${tier3YoY === null ? '' : `，同比 ${tier3YoY >= 0 ? '+' : ''}${formatPercent(tier3YoY)}`}`
          : '当前没有可用的三级医院趋势数据。',
        agingSeries: agingBuckets.map((label) => ({ label, value: agingMap[label] || 0 })),
        forecastRows: hospitals.filter((item) => item.unpaid > 0),
        riskRows: hospitals.filter((item) => item.unpaid > 0).slice(0, 5),
        selectedHospital,
        selectedHospitalCycles,
        hospitalOptions,
        summary: salesTotal
          ? `当前窗口内折后销售额 ${formatCurrency(salesTotal)}，回款率 ${formatPercent(collectionRate)}，未回款 ${formatCurrency(unpaidTotal)}。`
          : '当前筛选范围内没有可展示的销售与回款数据。',
        focus: topRisk && topRisk.unpaid > 0
          ? `重点关注 ${topRisk.hospital}，未回款 ${formatCurrency(topRisk.unpaid)}，最大账龄 ${formatNumber(topRisk.maxAging)} 天。`
          : '当前没有突出欠款医院，可继续保持常规回款节奏。',
        salesInsight: topHospitals.length
          ? `销量 Top1 医院是 ${topHospitals[0].hospital}，贡献 ${formatCurrency(topHospitals[0].sales)}。`
          : '当前没有可用销售医院数据。',
        hospitalInsight: strongestCity
          ? `${strongestCity[0]} 是当前主要销售贡献城市。`
          : '当前没有识别到主要销售城市。',
        projectInsight: topProject
          ? `${topProject.name} 是当前销量最高的项目，建议重点复盘其适配医院。`
          : '当前没有可用项目结构数据。',
        collectionInsight: unpaidTotal
          ? `未回款 ${formatCurrency(unpaidTotal)}，超信用期金额占比 ${formatPercent(overdueShare)}。`
          : '当前没有未回款压力。',
        forecastInsight: topRisk && topRisk.unpaid > 0
          ? `${topRisk.hospital} 在预测表中优先级最高，建议立即拉通回款动作。`
          : '当前预测表未识别出明显高风险医院。',
        agingInsight: overdueShare > 0
          ? `超信用期欠款占比 ${formatPercent(overdueShare)}，需重点压降高账龄医院。`
          : '当前欠款账龄结构较为平稳。',
        planTips: planTips.slice(0, 5),
      };
    }

    function populateStaticMeta() {
      document.getElementById('meta-range').textContent = `${RAW.defaultRange.from} 至 ${RAW.defaultRange.to}`;
      document.getElementById('meta-generated').textContent = RAW.generatedAt;
      const warnings = RAW.meta.warnings || [];
      const warningBox = document.getElementById('warning-box');
      if (warnings.length) {
        warningBox.style.display = 'block';
        warningBox.innerHTML = `<strong>数据提示</strong><br>${warnings.join('<br>')}`;
      }
    }

    function populateMonthSelectors() {
      const fromSel = document.getElementById('from-month');
      const toSel = document.getElementById('to-month');
      RAW.months.forEach((month) => {
        const a = document.createElement('option');
        a.value = month;
        a.textContent = month;
        fromSel.appendChild(a);
        const b = document.createElement('option');
        b.value = month;
        b.textContent = month;
        toSel.appendChild(b);
      });
      fromSel.value = state.from;
      toSel.value = state.to;
      fromSel.addEventListener('change', () => {
        state.from = fromSel.value;
        if (state.from > state.to) {
          state.to = state.from;
          toSel.value = state.to;
        }
        state.forecastExpanded = false;
        updateDashboard();
      });
      toSel.addEventListener('change', () => {
        state.to = toSel.value;
        if (state.to < state.from) {
          state.from = state.to;
          fromSel.value = state.from;
        }
        state.forecastExpanded = false;
        updateDashboard();
      });
    }

    function populateRegionSelector() {
      const regionSel = document.getElementById('region-filter');
      regionSel.innerHTML = '';
      ['全部', ...RAW.regions].forEach((region) => {
        const option = document.createElement('option');
        option.value = region;
        option.textContent = region;
        regionSel.appendChild(option);
      });
      regionSel.value = state.region;
      regionSel.addEventListener('change', () => {
        state.region = regionSel.value;
        state.city = '全部';
        state.forecastExpanded = false;
        updateDashboard();
      });
    }

    function refreshCitySelector() {
      const citySel = document.getElementById('city-filter');
      const cities = new Set();
      RAW.salesRecords.forEach((row) => {
        if ((state.region === '全部' || row.region === state.region) && row.city) cities.add(row.city);
      });
      RAW.arRecords.forEach((row) => {
        if ((state.region === '全部' || row.region === state.region) && row.city) cities.add(row.city);
      });
      const options = ['全部', ...[...cities].sort()];
      citySel.innerHTML = '';
      options.forEach((city) => {
        const option = document.createElement('option');
        option.value = city;
        option.textContent = city;
        citySel.appendChild(option);
      });
      if (!options.includes(state.city)) {
        state.city = '全部';
      }
      citySel.value = state.city;
      if (!citySel.dataset.bound) {
        citySel.addEventListener('change', () => {
          state.city = citySel.value;
          state.forecastExpanded = false;
          updateDashboard();
        });
        citySel.dataset.bound = '1';
      }
    }

    function refreshHospitalSelector(options) {
      const hospitalSel = document.getElementById('hospital-filter');
      hospitalSel.innerHTML = '';
      if (!options.length) {
        const option = document.createElement('option');
        option.value = '';
        option.textContent = '当前范围内无医院';
        hospitalSel.appendChild(option);
        state.hospital = '';
      } else {
        options.forEach((hospital) => {
          const option = document.createElement('option');
          option.value = hospital;
          option.textContent = hospital;
          hospitalSel.appendChild(option);
        });
        hospitalSel.value = state.hospital;
      }
      if (!hospitalSel.dataset.bound) {
        hospitalSel.addEventListener('change', () => {
          state.hospital = hospitalSel.value;
          updateDashboard();
        });
        hospitalSel.dataset.bound = '1';
      }
    }

    function bindForecastToggle() {
      const toggle = document.getElementById('forecast-toggle');
      toggle.addEventListener('click', () => {
        state.forecastExpanded = !state.forecastExpanded;
        updateDashboard();
      });
    }

    function buildCharts() {
      const endValueLabelPlugin = {
        id: 'endValueLabel',
        afterDatasetsDraw(chart, args, pluginOptions) {
          if (!pluginOptions || !pluginOptions.enabled) return;
          const datasetIndex = pluginOptions.datasetIndex || 0;
          const dataset = chart.data.datasets[datasetIndex];
          const meta = chart.getDatasetMeta(datasetIndex);
          if (!dataset || !meta || !meta.data || chart.config.options.indexAxis !== 'y') return;

          const { ctx, chartArea } = chart;
          ctx.save();
          ctx.fillStyle = pluginOptions.color || '#334e5c';
          ctx.font = '12px "Avenir Next","PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif';
          ctx.textBaseline = 'middle';

          meta.data.forEach((bar, index) => {
            const raw = Number(dataset.data[index] || 0);
            const label = pluginOptions.formatter ? pluginOptions.formatter(raw) : String(raw);
            const width = ctx.measureText(label).width;
            let x = bar.x + 8;
            let align = 'left';
            if (x + width > chartArea.right - 4) {
              x = chartArea.right - 4;
              align = 'right';
            }
            ctx.textAlign = align;
            ctx.fillText(label, x, bar.y);
          });
          ctx.restore();
        },
      };

      Chart.register(endValueLabelPlugin);
      Chart.defaults.color = '#5a707d';
      Chart.defaults.font.family = '"Avenir Next","PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif';
      Chart.defaults.plugins.tooltip.titleFont = { size: 13, weight: '700' };
      Chart.defaults.plugins.tooltip.bodyFont = { size: 12 };
      Chart.defaults.plugins.legend.labels.usePointStyle = true;
      charts.monthly = new Chart(document.getElementById('chart-monthly'), {
        type: 'bar',
        data: {
          labels: [],
          datasets: [
            {
              type: 'bar',
              label: '折后销售额',
              data: [],
              backgroundColor: 'rgba(15,118,110,.78)',
              borderRadius: 8,
              yAxisID: 'y',
              order: 2,
            },
            {
              type: 'line',
              label: '回款率',
              data: [],
              borderColor: '#2563eb',
              backgroundColor: '#2563eb',
              borderWidth: 2,
              pointRadius: 3,
              pointHoverRadius: 4,
              tension: .28,
              yAxisID: 'y1',
              order: 1,
            },
          ],
        },
        options: {
          maintainAspectRatio: false,
          layout: { padding: { top: 6, right: 8, bottom: 0, left: 0 } },
          plugins: {
            legend: { position: 'top', align: 'start' },
            tooltip: {
              callbacks: {
                label: (item) => item.dataset.label === '回款率'
                  ? `回款率 ${Number(item.raw || 0).toFixed(1)}%`
                  : `折后销售额 ${formatCurrency(item.raw)}`,
              },
            },
          },
          scales: {
            x: { grid: { display: false }, ticks: { maxRotation: 0, autoSkip: true, maxTicksLimit: 8 } },
            y: {
              beginAtZero: true,
              ticks: { callback: (value) => formatCurrency(value) },
              grid: { color: 'rgba(18,35,43,.07)' },
            },
            y1: {
              position: 'right',
              beginAtZero: true,
              suggestedMax: 100,
              ticks: { callback: (value) => formatPercentAxis(value) },
              grid: { drawOnChartArea: false },
            },
          },
        },
      });
      charts.topHospitals = new Chart(document.getElementById('chart-hospital-top10'), {
        type: 'bar',
        data: {
          labels: [],
          datasets: [{
            label: '折后销售额',
            data: [],
            backgroundColor: 'rgba(37,99,235,.78)',
            borderRadius: 10,
            barThickness: 18,
            maxBarThickness: 20,
          }],
        },
        options: {
          maintainAspectRatio: false,
          indexAxis: 'y',
          layout: { padding: { top: 4, right: 10, bottom: 4, left: 6 } },
          plugins: {
            legend: { display: false },
            endValueLabel: { enabled: true, formatter: formatCurrency },
            tooltip: {
              callbacks: {
                title: (items) => items[0] ? items[0].label : '',
                label: (item) => `折后销售额 ${formatCurrency(item.raw)}`,
              },
            },
          },
          scales: {
            x: {
              beginAtZero: true,
              grace: '22%',
              ticks: { callback: (value) => formatCurrency(value) },
              grid: { color: 'rgba(18,35,43,.07)' },
            },
            y: {
              grid: { display: false },
              ticks: {
                padding: 10,
                callback: function(value) { return categoryTickLabel(this, value, 12); },
              },
            },
          },
        },
      });
      charts.topProjects = new Chart(document.getElementById('chart-project-top10'), {
        type: 'bar',
        data: {
          labels: [],
          datasets: [{
            label: '折后销售额',
            data: [],
            backgroundColor: 'rgba(15,118,110,.72)',
            borderRadius: 10,
            barThickness: 18,
            maxBarThickness: 20,
          }],
        },
        options: {
          maintainAspectRatio: false,
          indexAxis: 'y',
          layout: { padding: { top: 4, right: 10, bottom: 4, left: 6 } },
          plugins: {
            legend: { display: false },
            endValueLabel: { enabled: true, formatter: formatCurrency },
            tooltip: {
              callbacks: {
                title: (items) => items[0] ? items[0].label : '',
                label: (item) => `折后销售额 ${formatCurrency(item.raw)}`,
              },
            },
          },
          scales: {
            x: {
              beginAtZero: true,
              grace: '22%',
              ticks: { callback: (value) => formatCurrency(value) },
              grid: { color: 'rgba(18,35,43,.07)' },
            },
            y: {
              grid: { display: false },
              ticks: {
                padding: 10,
                callback: function(value) { return categoryTickLabel(this, value, 12); },
              },
            },
          },
        },
      });
      charts.tier3 = new Chart(document.getElementById('chart-tier3-trend'), {
        type: 'line',
        data: { labels: [], datasets: [{ label: '折后销售额', data: [], borderColor: '#2563eb', backgroundColor: 'rgba(37,99,235,.12)', tension: .28, fill: true }] },
        options: {
          maintainAspectRatio: false,
          layout: { padding: { top: 8, right: 8, bottom: 0, left: 0 } },
          plugins: { legend: { display: false } },
          scales: {
            x: { grid: { display: false }, ticks: { maxRotation: 0, autoSkip: false, callback: function(value) { return quarterTickLabel(this, value); } } },
            y: { beginAtZero: true, ticks: { callback: (value) => formatCurrency(value) }, grid: { color: 'rgba(18,35,43,.07)' } },
          },
        },
      });
      charts.aging = new Chart(document.getElementById('chart-aging'), {
        type: 'bar',
        data: {
          labels: agingBuckets,
          datasets: [{
            data: [],
            backgroundColor: ['#92d2c6', '#56b7a7', '#3f7ae0', '#f59e0b', '#dc2626'],
            borderRadius: 10,
            maxBarThickness: 38,
          }],
        },
        options: {
          maintainAspectRatio: false,
          plugins: {
            legend: { display: false },
            tooltip: {
              callbacks: {
                label: (item) => `未回款 ${formatCurrency(item.raw)}`,
              },
            },
          },
          scales: {
            x: { grid: { display: false } },
            y: { beginAtZero: true, ticks: { callback: (value) => formatCurrency(value) }, grid: { color: 'rgba(18,35,43,.07)' } },
          },
        },
      });
      charts.hospitalCycle = new Chart(document.getElementById('chart-hospital-cycle'), {
        type: 'bar',
        data: {
          labels: [],
          datasets: [{
            label: '回款周期',
            data: [],
            backgroundColor: 'rgba(37,99,235,.78)',
            borderRadius: 8,
            maxBarThickness: 28,
          }],
        },
        options: {
          maintainAspectRatio: false,
          plugins: {
            legend: { display: false },
            tooltip: {
              callbacks: {
                label: (item) => `回款周期 ${formatNumber(item.raw)} 天`,
              },
            },
          },
          scales: {
            x: { grid: { display: false } },
            y: {
              beginAtZero: true,
              ticks: { callback: (value) => `${formatNumber(value)} 天` },
              grid: { color: 'rgba(18,35,43,.07)' },
            },
          },
        },
      });
    }

    function renderForecastTable(rows) {
      const tbody = document.querySelector('#forecast-table tbody');
      const toggle = document.getElementById('forecast-toggle');
      const summary = document.getElementById('forecast-table-summary');
      tbody.innerHTML = '';
      if (!rows.length) {
        toggle.style.display = 'none';
        summary.textContent = '当前筛选范围内没有医院回款预测数据。';
        const tr = document.createElement('tr');
        tr.innerHTML = '<td data-label="结果" colspan="9">当前筛选范围内没有医院回款预测数据。</td>';
        tbody.appendChild(tr);
        return;
      }
      const visibleRows = state.forecastExpanded ? rows : rows.slice(0, 8);
      toggle.style.display = rows.length > 8 ? 'inline-flex' : 'none';
      toggle.textContent = state.forecastExpanded ? '收起明细' : '展开全部';
      summary.textContent = rows.length > 8
        ? (state.forecastExpanded
          ? `当前已展开完整明细，共 ${rows.length} 家医院。`
          : `当前默认展示前 8 家医院，完整明细共 ${rows.length} 家。`)
        : `当前范围内共有 ${rows.length} 家医院进入预测表。`;

      visibleRows.forEach((row) => {
        const tr = document.createElement('tr');
        tr.innerHTML = `
          <td data-label="医院">${row.hospital}</td>
          <td data-label="区域 / 城市">${row.region} / ${row.city}</td>
          <td data-label="未回款">${formatCurrency(row.unpaid)}</td>
          <td data-label="平均回款周期">${row.avgCycle ? `${formatNumber(row.avgCycle)} 天` : '-'}</td>
          <td data-label="最近回款周期">${row.latestCycle ? `${formatNumber(row.latestCycle)} 天` : '-'}</td>
          <td data-label="信用天数">${formatNumber(row.creditDays)} 天</td>
          <td data-label="预计回款日期">${row.predictedDate || '-'}</td>
          <td data-label="风险等级"><span class="badge ${badgeClass(row.riskLevel)}">${row.riskLevel}</span></td>
          <td data-label="建议动作">${row.suggestion}</td>
        `;
        tbody.appendChild(tr);
      });
    }

    function renderRiskTable(rows) {
      const tbody = document.querySelector('#risk-table tbody');
      tbody.innerHTML = '';
      if (!rows.length) {
        const tr = document.createElement('tr');
        tr.innerHTML = '<td data-label="结果" colspan="4">当前没有欠款医院。</td>';
        tbody.appendChild(tr);
        return;
      }
      rows.forEach((row) => {
        const tr = document.createElement('tr');
        tr.innerHTML = `
          <td data-label="医院">${row.hospital}</td>
          <td data-label="未回款">${formatCurrency(row.unpaid)}</td>
          <td data-label="最大账龄">${formatNumber(row.maxAging)} 天</td>
          <td data-label="风险"><span class="badge ${badgeClass(row.riskLevel)}">${row.riskLevel}</span></td>
        `;
        tbody.appendChild(tr);
      });
    }

    function renderHospitalDetail(entry) {
      const grid = document.getElementById('hospital-detail-grid');
      const caption = document.getElementById('hospital-cycle-caption');
      grid.innerHTML = '';
      if (!entry) {
        grid.innerHTML = '<div class="detail-box"><span>提示</span><strong>当前筛选范围内没有可展示的医院回款明细。</strong></div>';
        caption.textContent = '没有可用于展示的回款周期记录。';
        return;
      }
      const items = [
        ['医院', entry.hospital],
        ['未回款', formatCurrency(entry.unpaid)],
        ['平均回款周期', entry.avgCycle ? `${formatNumber(entry.avgCycle)} 天` : '-'],
        ['最近回款周期', entry.latestCycle ? `${formatNumber(entry.latestCycle)} 天` : '-'],
        ['信用天数', `${formatNumber(entry.creditDays)} 天`],
      ];
      items.forEach(([label, value]) => {
        const el = document.createElement('div');
        el.className = 'detail-box';
        el.innerHTML = `<span>${label}</span><strong>${value}</strong>`;
        grid.appendChild(el);
      });
      caption.textContent = entry.recentCycleItems.length
        ? `${entry.hospital} 当前为 ${entry.riskLevel}，预计回款日期 ${entry.predictedDate || '-'}，加权账龄 ${formatNumber(entry.weightedAging)} 天。`
        : `${entry.hospital} 暂无历史回款周期记录，当前按信用天数做保守预测。`;
    }

    function renderTips(tips) {
      const container = document.getElementById('tips-list');
      container.innerHTML = '';
      if (!tips.length) {
        const tip = document.createElement('div');
        tip.className = 'tip';
        tip.textContent = '当前筛选范围内暂未生成计划提示，请调整时间、区域或城市后查看。';
        container.appendChild(tip);
        return;
      }
      tips.forEach((text, index) => {
        const item = document.createElement('div');
        item.className = 'tip';
        item.innerHTML = `<strong>建议 ${index + 1}</strong><br>${text}`;
        container.appendChild(item);
      });
    }

    function updateDashboard() {
      refreshCitySelector();
      const aggregate = buildAggregate();
      refreshHospitalSelector(aggregate.hospitalOptions);

      animateValue(document.getElementById('kpi-sales'), formatCurrency, aggregate.metrics.salesTotal);
      animateValue(document.getElementById('kpi-specimen'), (v) => formatCompact(v, '份'), aggregate.metrics.specimenTotal);
      animateValue(document.getElementById('kpi-unpaid'), formatCurrency, aggregate.metrics.unpaidTotal);
      animateValue(document.getElementById('kpi-rate'), (v) => `${v.toFixed(1)}%`, aggregate.metrics.collectionRate * 100);

      document.getElementById('summary-text').textContent = aggregate.summary;
      document.getElementById('sales-tag').textContent = `${state.from} 至 ${state.to} / ${state.region}${state.city === '全部' ? '' : ` / ${state.city}`}`;
      document.getElementById('collection-tag').textContent = `${state.region}${state.city === '全部' ? '' : ` / ${state.city}`} / 风险优先`;
      document.getElementById('sales-insight').textContent = aggregate.salesInsight;
      document.getElementById('hospital-insight').textContent = aggregate.hospitalInsight;
      document.getElementById('project-insight').textContent = aggregate.projectInsight;
      document.getElementById('tier3-summary').textContent = aggregate.tier3Summary;
      document.getElementById('collection-insight').textContent = aggregate.collectionInsight;
      document.getElementById('forecast-insight').textContent = aggregate.forecastInsight;
      document.getElementById('aging-insight').textContent = aggregate.agingInsight;
      document.getElementById('kpi-sales-hint').textContent = `当前窗口内共有 ${formatCompact(aggregate.metrics.specimenTotal, '份')} 标本量，可用于判断临床上量强弱。`;
      document.getElementById('kpi-specimen-hint').textContent = `时间窗口覆盖 ${aggregate.months.length} 个自然月，适合做周更和月末复盘。`;
      document.getElementById('kpi-unpaid-hint').textContent = `超信用期欠款占比 ${formatPercent(aggregate.metrics.overdueShare)}。`;
      document.getElementById('kpi-rate-hint').textContent = `当前窗口回款率 ${formatPercent(aggregate.metrics.collectionRate)}，建议结合预测表同步看重点医院。`;

      charts.monthly.data.labels = aggregate.monthlySales.map((item) => item.month);
      charts.monthly.data.datasets[0].data = aggregate.monthlySales.map((item) => item.value);
      charts.monthly.data.datasets[1].data = aggregate.monthlyCollectionRate.map((item) => Number((item.value * 100).toFixed(1)));
      charts.monthly.update();

      charts.topHospitals.data.labels = aggregate.topHospitals.map((item) => item.hospital);
      charts.topHospitals.data.datasets[0].data = aggregate.topHospitals.map((item) => item.sales);
      charts.topHospitals.update();

      charts.topProjects.data.labels = aggregate.topProjects.map((item) => item.name);
      charts.topProjects.data.datasets[0].data = aggregate.topProjects.map((item) => item.value);
      charts.topProjects.update();

      charts.tier3.data.labels = aggregate.tier3Trend.map((item) => item.month);
      charts.tier3.data.datasets[0].data = aggregate.tier3Trend.map((item) => item.value);
      charts.tier3.update();

      charts.aging.data.datasets[0].data = aggregate.agingSeries.map((item) => item.value);
      charts.aging.update();

      charts.hospitalCycle.data.labels = aggregate.selectedHospitalCycles.map((item) => item.label);
      charts.hospitalCycle.data.datasets[0].data = aggregate.selectedHospitalCycles.map((item) => item.value);
      charts.hospitalCycle.update();

      renderForecastTable(aggregate.forecastRows);
      renderRiskTable(aggregate.riskRows);
      renderHospitalDetail(aggregate.selectedHospital);
      renderTips(aggregate.planTips);
    }

    function observeEntrance() {
      const observer = new IntersectionObserver((entries) => {
        entries.forEach((entry) => {
          if (entry.isIntersecting) {
            entry.target.classList.add('visible');
            observer.unobserve(entry.target);
          }
        });
      }, { threshold: 0.12 });
      document.querySelectorAll('.fade-up').forEach((node) => observer.observe(node));
    }

    populateStaticMeta();
    populateMonthSelectors();
    populateRegionSelector();
    bindForecastToggle();
    buildCharts();
    updateDashboard();
    observeEntrance();
  </script>
</body>
</html>
"""
    return template.replace("__CHART_JS__", chart_js).replace("__DATA_JSON__", payload_json)


def print_summary(payload: dict[str, object], config: RuntimeConfig, warnings: list[str]) -> None:
    snapshot = payload["meta"]["snapshot"]
    default_range = payload["defaultRange"]
    print("Dashboard V3 refreshed successfully")
    print(f"Output HTML: {config.html_out_path}")
    print(f"Default range: {default_range['from']} -> {default_range['to']}")
    print(f"Sales rows: {payload['meta']['sales']['row_count']}")
    print(f"AR rows: {payload['meta']['ar']['row_count']}")
    print(f"Sales total (12m): {snapshot['sales_total']:.2f}")
    print(f"Unpaid total (12m): {snapshot['unpaid_total']:.2f}")
    print(f"Collection rate (12m): {snapshot['collection_rate'] * 100:.2f}%")
    print(f"Sales source: {config.sales_path}")
    print(f"AR source: {config.ar_path}")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"  - {warning}")


def main() -> None:
    config = parse_args()
    warnings: list[str] = []
    config.html_out_path.parent.mkdir(parents=True, exist_ok=True)

    sales_records, sales_meta = read_sales_records(config.sales_path, warnings)
    ar_records, ar_meta = read_ar_records(config.ar_path, warnings)
    ar_records = annotate_ar_records(sales_records, ar_records)

    payload = build_payload(sales_records, ar_records, sales_meta, ar_meta, config, warnings)
    html = render_html(payload, config.chart_js_path)
    config.html_out_path.write_text(html, encoding="utf-8")
    print_summary(payload, config, warnings)


if __name__ == "__main__":
    try:
        main()
    except ValidationError as exc:
        raise SystemExit(f"ERROR: {exc}")
